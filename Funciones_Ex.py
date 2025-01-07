"""
Este módulo contiene funciones y clases para interactuar con archivos 
Excel utilizando la librería openpyxl.
Las funciones permiten leer y escribir datos, así como obtener la cantidad de 
filas y columnas en una hoja de Excel.

Funciones:
- getRowCount: Obtiene el número de filas en una hoja de Excel.
- getColumnCount: Obtiene el número de columnas en una hoja de Excel.
- readData: Lee datos de una celda específica en una hoja de Excel.
- writeData: Escribe datos en una celda específica en una hoja de Excel.
"""

import openpyxl
from openpyxl.workbook import Workbook
import xlwings as xw


class Funexcel():
    """
    Esta clase proporciona métodos para interactuar con archivos Excel utilizando la biblioteca openpyxl.
    """

    def get_row_count(self, path, sheetName):
        """
        Obtiene el número de filas en una hoja de Excel.
        :param path: Ruta del archivo de Excel
        :param sheet_name: Nombre de la hoja de Excel
        :return: Número de filas en la hoja
        """
        Workbook = openpyxl.load_workbook(path)
        sheet = Workbook[sheetName]
        return (sheet.max_row)


    def get_column_count(self, file, sheetName):
        """
        Obtiene el número de columnas en una hoja de Excel.
        :param path: Ruta del archivo de Excel
        :param sheet_name: Nombre de la hoja de Excel
        :return: Número de columnas en la hoja
        """
        Workbook = openpyxl.load_workbook(file)
        sheet = Workbook[sheetName]
        return (sheet.max_column)

    def read_data(self, path, sheetName, rownum, columna):
        """
        Lee datos de una celda en una hoja de Excel.
        :param path: Ruta del archivo de Excel
        :param sheet_name: Nombre de la hoja de Excel
        :param rownum: Número de fila
        :param column: Número de columna
        :return: El valor de la celda especificada
        """
        Workbook = openpyxl.load_workbook(path)
        sheet = Workbook[sheetName]
        return sheet.cell(row=rownum, column=columna).value
    
    def read_data2(self, path, sheetName, rownum, columna):
        """
        Lee datos de una celda en una hoja de Excel, evaluando fórmulas.
        :param path: Ruta del archivo de Excel
        :param sheet_name: Nombre de la hoja de Excel
        :param rownum: Número de fila
        :param columna: Número de columna
        :return: El valor de la celda especificada
        """
        # Abrir el libro con xlwings
        with xw.App(visible=False) as app:
            wb = app.books.open(path)
            sheet = wb.sheets[sheetName]
            # Obtener el valor de la celda, evaluando fórmulas
            return sheet.cells(rownum, columna).value

    def write_data(self, file, path, sheetName, rownum, columna, data):
        """
        Escribe datos en una celda en una hoja de Excel.
        :param path: Ruta del archivo de Excel
        :param sheet_name: Nombre de la hoja de Excel
        :param rownum: Número de fila
        :param column: Número de columna
        :param data: Datos a escribir en la celda
        """
        Workbook = openpyxl.load_workbook(path)
        sheet = Workbook[sheetName]
        sheet.cell(row=rownum, column=columna).value = data
        Workbook.save(path)