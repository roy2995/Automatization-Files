import os
from docxtpl import DocxTemplate
from src.utils import clean_filename, split_steps
from openpyxl import load_workbook

def main():
    # Configuración del archivo Excel, hoja, y rango
    excel_file = "config/entregable5_v2.xlsx"  
    sheet_name = "MP"
    start_row = 1  
    column_map = {
        "name": 1,      
        "title": 7,     
        "steps": 9,      
        "platform": 4,  
        "id": 3,
        "module": 5
    }

    try:
        # Cargar el archivo Excel con valores calculados (no fórmulas)
        workbook = load_workbook(excel_file, data_only=True)
        sheet = workbook[sheet_name]

        # Crear la carpeta "Files" si no existe
        output_dir = "Files"
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        # Iterar sobre las filas desde la fila inicial hasta el final
        for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, values_only=True):
            # Extraer valores usando el mapa de columnas
            name = row[column_map["name"] - 1]
            title = row[column_map["title"] - 1]
            steps = row[column_map["steps"] - 1]
            platform = row[column_map["platform"] - 1]
            _id = row[column_map["id"] - 1]
            module = row[column_map["module"] - 1]

            # Saltar filas incompletas
            if not all([name, title, steps, platform, _id]):
                print(f"Fila incompleta, saltando: {row}")
                continue

            # Dividir los pasos de prueba (si están en una sola celda)
            pasos = split_steps(steps)

            # Preparar el contexto para la plantilla
            context = {
                "name": name,
                "title": title,
                "paso_prueba": pasos,
                "platform": platform,
                "id": _id,
                "module": module
            }

            # Cargar la plantilla de Word
            template_path = "TestFile.docx"  # Ruta de la plantilla
            if not os.path.exists(template_path):
                print(f"Plantilla no encontrada en {template_path}. Asegúrate de que existe.")
                return

            doc = DocxTemplate(template_path)

            # Rellenar la plantilla con los datos de la fila actual
            doc.render(context)

            # Generar un nombre de archivo limpio
            clean_title = clean_filename(f"FormatoDeEvidencia_CP_{_id}_{platform}")
            output_filename = os.path.join(output_dir, f"{clean_title}.docx")

            # Guardar el archivo Word
            try:
                doc.save(output_filename)
                print(f"Archivo generado: {output_filename}")
            except Exception as e:
                print(f"Error al guardar el archivo Word: {e}")

    except PermissionError as e:
        print(f"Error de permisos al leer el archivo Excel: {e}")
    except FileNotFoundError as e:
        print(f"Archivo Excel no encontrado: {e}")
    except Exception as e:
        print(f"Error al procesar el archivo Excel: {e}")

if __name__ == "__main__":
    main()
